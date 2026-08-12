#!/usr/bin/env python3
"""External scenario emitter — NOT a Hermes tool.

Compresses Kaggle JPC Daily Site Diary / Work Plan / EHS forms + Tasks into a
7-day wall-clock email campaign. Emails are sent (or dry-run printed) as if
from site staff / EHS / suppliers. Hermes only sees normal IMAP mail and must
act naturally — it is never told this is a simulation.

Modes:
  dry-run  — print schedule (default)
  smtp     — send via SMTP to HERMES_INBOX (real Gmail → Hermes email platform)

Env:
  FIELDCLAW_KAGGLE_DIR   default: <repo>/research/site-logs/kaggle-jpc
  HERMES_INBOX           recipient (Hermes EMAIL_ADDRESS)
  SMTP_HOST/PORT/USER/PASSWORD/FROM
  WEEK_LOOP_SPEED        wall seconds per scenario-day (default 86400; use 60 for demo)
  WEEK_LOOP_STATE        path to state json (resume cursor)
"""

from __future__ import annotations

import argparse
import json
import os
import smtplib
import time
from collections import defaultdict
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("openpyxl required: uv pip install openpyxl") from e

REPO = Path(__file__).resolve().parents[2]
DEFAULT_KAGGLE = REPO / "research" / "site-logs" / "kaggle-jpc"


def _as_dt(v):
    return v if isinstance(v, datetime) else None


def load_kaggle_events(kaggle_dir: Path) -> list[dict]:
    forms = list(
        openpyxl.load_workbook(kaggle_dir / "Forms.xlsx", read_only=True, data_only=True)
        .active.iter_rows(values_only=True)
    )
    tasks = list(
        openpyxl.load_workbook(kaggle_dir / "Tasks.xlsx", read_only=True, data_only=True)
        .active.iter_rows(values_only=True)
    )
    fh, frows = list(forms[0]), forms[1:]
    th, trows = list(tasks[0]), tasks[1:]
    fi = {k: fh.index(k) for k in ("Location", "Type", "Created", "Name", "Status", "Project")}
    ti = {
        k: th.index(k)
        for k in (
            "Location",
            "Type",
            "Created",
            "Description",
            "Status",
            "project",
            "Priority",
            "Task Group",
        )
    }

    def is_site_form(loc, typ):
        s = f"{loc or ''} {typ or ''}".lower()
        return any(
            k in s
            for k in (
                "daily site diary",
                "daily work plan",
                "ehs inspection",
                "ehs inspections",
                "ehs forms",
            )
        )

    by_day: dict = defaultdict(list)
    for r in frows:
        if not is_site_form(r[fi["Location"]], r[fi["Type"]]):
            continue
        d = _as_dt(r[fi["Created"]])
        if not d:
            continue
        by_day[d.date()].append(
            {
                "kind": "form",
                "created": d,
                "subject": f"[Site] {r[fi['Name']] or 'Daily report'} — Project {r[fi['Project']]}",
                "body": (
                    f"Site report filed.\n\n"
                    f"Form: {r[fi['Name']]}\n"
                    f"Type: {r[fi['Type']]}\n"
                    f"Location: {r[fi['Location']]}\n"
                    f"Status: {r[fi['Status']]}\n"
                    f"Project: {r[fi['Project']]}\n"
                    f"Filed: {d.isoformat()}\n\n"
                    f"— Site management (automated field capture)\n"
                ),
                "from_role": "site.management",
            }
        )

    active_days = [d for d in sorted(by_day) if by_day[d]][-21:]  # ~3 weeks source → 7 day loop
    if not active_days:
        raise SystemExit("no diary forms found in Kaggle data")

    # Pack source days into 7 scenario days
    buckets: list[list[dict]] = [[] for _ in range(7)]
    for i, day in enumerate(active_days):
        buckets[i % 7].extend(by_day[day][:4])  # cap per source day

    # Add safety/quality tasks into same window
    d0, d1 = active_days[0], active_days[-1]
    task_i = 0
    for r in trows:
        d = _as_dt(r[ti["Created"]])
        if not d or not (d0 <= d.date() <= d1):
            continue
        typ = str(r[ti["Type"]] or "")
        group = str(r[ti["Task Group"]] or "")
        if not any(x in typ or x == group for x in ("Safety", "Quality", "Snag", "Defect")):
            continue
        desc = r[ti["Description"]] or typ
        if isinstance(desc, str) and len(desc) > 400:
            desc = desc[:400] + "…"
        role = "ehs" if "Safety" in typ or group == "Safety" else "quality"
        buckets[task_i % 7].append(
            {
                "kind": "task",
                "created": d,
                "subject": f"[{role.upper()}] {typ}",
                "body": (
                    f"{desc}\n\n"
                    f"Type: {typ}\n"
                    f"Priority: {r[ti['Priority']]}\n"
                    f"Location: {r[ti['Location']]}\n"
                    f"Status: {r[ti['Status']]}\n"
                    f"Project: {r[ti['project']]}\n"
                    f"Raised: {d.isoformat()}\n\n"
                    f"— {role} desk\n"
                ),
                "from_role": role,
            }
        )
        task_i += 1
        if task_i >= 40:
            break

    schedule = []
    for day_idx, items in enumerate(buckets):
        # Spread within the scenario-day
        n = max(len(items), 1)
        for j, item in enumerate(items):
            frac = (j + 0.5) / n  # within day
            schedule.append(
                {
                    "scenario_day": day_idx,  # 0..6
                    "frac": frac,
                    **item,
                }
            )
    schedule.sort(key=lambda x: (x["scenario_day"], x["frac"]))
    return schedule


FROM_MAP = {
    "site.management": "site-desk@fieldclaw.demo",
    "ehs": "ehs@fieldclaw.demo",
    "quality": "qa@fieldclaw.demo",
}


def _load_sim_env() -> None:
    """Load sim/.env.sim into os.environ (does not touch Hermes EMAIL_*)."""
    env_path = REPO / "sim" / ".env.sim"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.lstrip().startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k, v = k.strip(), v.strip().strip('"').strip("'")
        os.environ.setdefault(k, v)


def send_smtp(item: dict, to_addr: str) -> None:
    # Prefer SIM_* (Gmail sender). Legacy SMTP_* still accepted.
    host = os.environ.get("SIM_SMTP_HOST") or os.environ.get("SMTP_HOST", "smtp.gmail.com")
    port = int(os.environ.get("SIM_SMTP_PORT") or os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SIM_SMTP_USER") or os.environ["SMTP_USER"]
    password = os.environ.get("SIM_SMTP_PASSWORD") or os.environ["SMTP_PASSWORD"]
    from_addr = os.environ.get("SIM_SMTP_FROM") or os.environ.get("SMTP_FROM") or user

    msg = EmailMessage()
    msg["From"] = f"{item['from_role']} <{from_addr}>"
    msg["To"] = to_addr
    msg["Subject"] = item["subject"]
    msg["X-FieldClaw-Scenario-Day"] = str(item["scenario_day"])
    msg["X-FieldClaw-Source"] = "kaggle-jpc-week-loop"
    # Spoof Reply-Path style note in body only — From uses real SMTP user for deliverability
    body = (
        f"(Field capture from {FROM_MAP.get(item['from_role'], item['from_role'])})\n\n"
        + item["body"]
    )
    msg.set_content(body)

    with smtplib.SMTP(host, port, timeout=60) as smtp:
        smtp.starttls()
        smtp.login(user, password)
        smtp.send_message(msg)


def run_loop(schedule: list[dict], *, mode: str, speed: float, state_path: Path, once_day: int | None):
    state = {"sent": 0, "started_at": None}
    if state_path.exists():
        state.update(json.loads(state_path.read_text(encoding="utf-8")))

    inbox = os.environ.get("HERMES_INBOX") or os.environ.get("EMAIL_ADDRESS")
    if mode == "smtp" and not inbox:
        raise SystemExit("HERMES_INBOX or EMAIL_ADDRESS required for smtp mode")

    if state.get("started_at") is None:
        state["started_at"] = datetime.now(timezone.utc).isoformat()
        state_path.parent.mkdir(parents=True, exist_ok=True)
        state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")

    started = datetime.fromisoformat(state["started_at"])
    sent = int(state.get("sent", 0))

    print(f"week_email_loop mode={mode} speed={speed}s/day items={len(schedule)} sent={sent}")

    while sent < len(schedule):
        item = schedule[sent]
        if once_day is not None and item["scenario_day"] != once_day:
            sent += 1
            continue

        target_offset = (item["scenario_day"] + item["frac"]) * speed
        elapsed = (datetime.now(timezone.utc) - started).total_seconds()
        wait = target_offset - elapsed
        if wait > 0 and mode != "dump":
            print(f"sleep {wait:.1f}s until day={item['scenario_day']} #{sent}: {item['subject'][:60]}")
            time.sleep(min(wait, 3600))
            continue

        if mode == "dry-run" or mode == "dump":
            print(
                json.dumps(
                    {
                        "i": sent,
                        "scenario_day": item["scenario_day"],
                        "subject": item["subject"],
                        "from_role": item["from_role"],
                        "when_offset_sec": target_offset,
                    },
                    ensure_ascii=False,
                )
            )
        elif mode == "smtp":
            send_smtp(item, inbox)
            print(f"sent day={item['scenario_day']} #{sent}: {item['subject'][:70]}")
        else:
            raise SystemExit(f"unknown mode {mode}")

        sent += 1
        state["sent"] = sent
        state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")

        if mode == "dump":
            continue
        # small gap so IMAP sees distinct messages
        time.sleep(0.5)

    print("week loop complete" if once_day is None else f"day {once_day} dump/send done")


def main():
    _load_sim_env()
    p = argparse.ArgumentParser(description="7-day Kaggle→email scenario emitter (external)")
    p.add_argument("--mode", choices=("dry-run", "smtp", "dump"), default="dry-run")
    p.add_argument("--speed", type=float, default=float(os.environ.get("WEEK_LOOP_SPEED", "86400")))
    p.add_argument("--kaggle-dir", type=Path, default=Path(os.environ.get("FIELDCLAW_KAGGLE_DIR", DEFAULT_KAGGLE)))
    p.add_argument(
        "--state",
        type=Path,
        default=Path(os.environ.get("WEEK_LOOP_STATE", REPO / "data" / "week_loop_state.json")),
    )
    p.add_argument("--reset-state", action="store_true")
    p.add_argument("--day", type=int, default=None, help="Only emit one scenario day 0..6")
    args = p.parse_args()

    if args.reset_state and args.state.exists():
        args.state.unlink()

    schedule = load_kaggle_events(args.kaggle_dir)
    run_loop(schedule, mode=args.mode, speed=args.speed, state_path=args.state, once_day=args.day)


if __name__ == "__main__":
    main()
